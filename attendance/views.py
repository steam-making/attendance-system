from django.shortcuts import render, get_object_or_404, redirect
from .models import Student, Attendance, Setting, AttendanceSession
from django.utils import timezone
from django.contrib import messages
from .forms import StudentForm
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
import openpyxl
from collections import defaultdict, OrderedDict
from .forms import SchoolForm, SchoolSmsSettingsForm
from django.contrib.auth.decorators import login_required
from attendance.models import School
from .sms import resolve_and_render_message
import json
from datetime import date, datetime, timedelta

color_map_header = {
    "1부": "table-primary",
    "2부": "table-success",
    "3부": "table-warning",
    "4부": "table-info",
    "5부": "table-light",
}

school_colors = [
    "#e6f2ff",  # 연한 파랑
    "#e8f7e4",  # 연한 초록
    "#fff9e6",  # 연한 노랑
    "#f0f4ff",  # 연한 하늘
    "#f9f9f9",  # 연한 회색
]

@csrf_exempt
def update_today_attendance_status(request, student_id):
    if request.method == "PATCH":
        # (This function seems to have been broken, restoring original if possible or fixing syntax)
        # Assuming it was a simple status update API
        try:
            data = json.loads(request.body)
            new_status = data.get('status')
            attendance = get_object_or_404(Attendance, student_id=student_id, date=timezone.localdate())
            attendance.status = new_status
            attendance.save()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
def mark_attendance_end(request, student_id):
    if request.method == 'POST':
        student = get_object_or_404(Student, id=student_id, school__user=request.user)
        
        # ✅ Get date from request body if available, else default to today
        data = json.loads(request.body or "{}")
        date_str = data.get('date')
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                target_date = timezone.localdate()
        else:
            target_date = timezone.localdate()

        try:
            attendance = Attendance.objects.get(student=student, date=target_date)
            user_settings, _ = Setting.objects.get_or_create(user=student.school.user)
            
            status_to_save = '종료처리'
            if not user_settings.send_class_end_sms:
                status_to_save = '종료처리(문자x)'
                
            attendance.status = status_to_save
            attendance.save()
            return JsonResponse({
                'status': 'ended',
                'student': student.name,
                'phone': student.phone,
                'student_id': student.id,
                'final_status': status_to_save
            })
        except Attendance.DoesNotExist:
            return JsonResponse({'status': 'not_found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def update_school(request, pk):
    school = get_object_or_404(School, pk=pk, user=request.user)
    if request.method == 'POST':
        form = SchoolForm(request.POST, instance=school)
        if form.is_valid():
            form.save()
            return redirect('select_school')
    else:
        form = SchoolForm(instance=school)
    return render(request, 'attendance/register_school.html', {'form': form})


@login_required
def school_sms_settings(request, pk):
    school = get_object_or_404(School, pk=pk, user=request.user)
    user_settings, _ = Setting.objects.get_or_create(user=request.user)

    if request.method == 'POST' and request.POST.get('action') == 'reset_defaults':
        school.attendance_message_override = None
        school.lateness_message_override = None
        school.absence_message_override = None
        school.class_end_message_override = None
        school.cancel_message_override = None
        school.save(update_fields=[
            'attendance_message_override',
            'lateness_message_override',
            'absence_message_override',
            'class_end_message_override',
            'cancel_message_override',
        ])
        messages.success(request, f"{school.name} 학교의 문자 설정이 기본값으로 초기화되었습니다.")
        return redirect('school_sms_settings', pk=school.id)

    if request.method == 'POST':
        form = SchoolSmsSettingsForm(request.POST, instance=school, default_settings=user_settings)
        if form.is_valid():
            form.save()
            messages.success(request, f"{school.name} 학교 문자 설정이 저장되었습니다.")
            return redirect('school_sms_settings', pk=school.id)
    else:
        form = SchoolSmsSettingsForm(instance=school, default_settings=user_settings)

    return render(request, 'attendance/school_sms_settings.html', {
        'school': school,
        'form': form,
    })

@login_required
def delete_school(request, pk):
    school = get_object_or_404(School, pk=pk, user=request.user)
    if request.method == 'POST':
        school.delete()
        return redirect('select_school')
    return HttpResponse("허용되지 않은 접근입니다.", status=405)

@login_required
def move_student_department(request, pk):
    student = get_object_or_404(Student, pk=pk, school__user=request.user)

    if request.method == 'POST':
        new_department = request.POST.get('department')
        if new_department:
            student.department = new_department
            student.save()
            messages.success(request, f"{student.name} 학생이 {new_department}로 이동되었습니다.")
        return redirect(f"/attendance/list/?school={student.school.id}")

    return HttpResponse("허용되지 않은 접근입니다.", status=405)


def _weekday_label(target_date):
    labels = ['월', '화', '수', '목', '금', '토', '일']
    return labels[target_date.weekday()]


def _parse_time(value):
    if not value:
        return None
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    return None


@login_required
@require_POST
def start_class_session(request):
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '잘못된 데이터 형식입니다.'}, status=400)

    school_id = data.get('school_id')
    if not school_id:
        return JsonResponse({'status': 'error', 'message': 'school_id가 필요합니다.'}, status=400)

    school = get_object_or_404(School, id=school_id, user=request.user)
    today = timezone.localdate()
    now = timezone.localtime()

    session, _ = AttendanceSession.objects.get_or_create(school=school, date=today)
    session.started_at = now
    session.is_active = True
    session.save(update_fields=['started_at', 'is_active'])

    students = Student.objects.filter(school=school)
    student_ids = list(students.values_list('id', flat=True))
    existing_ids = set(
        Attendance.objects.filter(date=today, student__school=school)
        .values_list('student_id', flat=True)
    )
    missing_ids = [student_id for student_id in student_ids if student_id not in existing_ids]
    Attendance.objects.bulk_create(
        [Attendance(student_id=student_id, date=today, status='대기') for student_id in missing_ids]
    )

    return JsonResponse({
        'status': 'success',
        'started_at': now.strftime('%H:%M:%S'),
        'created': len(missing_ids)
    })


@login_required
@require_POST
def stop_class_session(request):
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '잘못된 데이터 형식입니다.'}, status=400)

    school_id = data.get('school_id')
    if not school_id:
        return JsonResponse({'status': 'error', 'message': 'school_id가 필요합니다.'}, status=400)

    school = get_object_or_404(School, id=school_id, user=request.user)
    today = timezone.localdate()

    AttendanceSession.objects.filter(school=school, date=today, is_active=True).update(is_active=False)

    return JsonResponse({
        'status': 'success',
        'is_active': False
    })


@login_required
@require_POST
def auto_process_attendance(request):
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '잘못된 데이터 형식입니다.'}, status=400)

    school_id = data.get('school_id')
    if not school_id:
        return JsonResponse({'status': 'error', 'message': 'school_id가 필요합니다.'}, status=400)

    school = get_object_or_404(School, id=school_id, user=request.user)
    today = timezone.localdate()
    now = timezone.localtime()

    session = AttendanceSession.objects.filter(school=school, date=today, is_active=True).first()
    if not session or not session.started_at:
        return JsonResponse({'status': 'not_started'})

    if now < session.started_at:
        return JsonResponse({'status': 'not_started'})

    class_days = []
    if school.class_days:
        class_days = [day.strip() for day in school.class_days.split(',') if day.strip()]
    if class_days and _weekday_label(today) not in class_days:
        return JsonResponse({'status': 'not_class_day'})

    settings, _ = Setting.objects.get_or_create(user=request.user)
    department_times = school.get_today_department_times(today)
    tz = timezone.get_current_timezone()

    lateness_results = {}
    end_results = {}

    for department, time_info in department_times.items():
        if not time_info:
            continue
        start_time = _parse_time(time_info.get('start'))
        end_time = _parse_time(time_info.get('end'))
        if not start_time or not end_time:
            continue

        start_dt = timezone.make_aware(datetime.combine(today, start_time), tz)
        end_dt = timezone.make_aware(datetime.combine(today, end_time), tz)

        # ✅ 자동화 기능은 Pro 전용
        if request.user.can_use_automation and settings.auto_send_lateness_sms and now >= (start_dt + timedelta(minutes=10)):
            students = Student.objects.filter(school=school, department=department)
            student_ids = list(students.values_list('id', flat=True))
            if student_ids:
                existing_qs = Attendance.objects.filter(date=today, student_id__in=student_ids)
                existing_ids = set(existing_qs.values_list('student_id', flat=True))
                updated_count = existing_qs.filter(status__in=['대기', '취소']).update(status='지각')
                missing_ids = [student_id for student_id in student_ids if student_id not in existing_ids]
                Attendance.objects.bulk_create(
                    [Attendance(student_id=student_id, date=today, status='지각') for student_id in missing_ids]
                )
                lateness_results[department] = {
                    'updated': updated_count,
                    'created': len(missing_ids)
                }

        # ✅ 자동화 기능은 Pro 전용
        if request.user.can_use_automation and settings.auto_send_class_end_sms and now >= end_dt:
            ended_count = Attendance.objects.filter(
                date=today,
                student__school=school,
                student__department=department,
                status='출석'
            ).update(status='종료처리')
            end_results[department] = {'ended': ended_count}

    return JsonResponse({
        'status': 'success',
        'lateness': lateness_results,
        'end': end_results
    })


@login_required
def delete_student(request, pk):
    student = get_object_or_404(Student, pk=pk, school__user=request.user)
    if request.method == 'POST':
        student.delete()
        return redirect(f"/attendance/list/?school={student.school.id}")
    return HttpResponse("허용되지 않은 접근입니다.", status=405)

@csrf_exempt
def delete_selected_students(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        Student.objects.filter(id__in=student_ids).delete()
        return JsonResponse({'status': 'success', 'deleted_count': len(student_ids)})

    return JsonResponse({'status': 'invalid_method'})

@login_required
def register_student(request):
    school_id = request.GET.get('school')  # ✅ URL에서 school ID 가져오기
    selected_school = School.objects.filter(id=school_id, user=request.user).first()

    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            student.school = selected_school  # ✅ 자동으로 학교 지정
            student.save()
            return redirect(f'/attendance/?school={selected_school.id}')
    else:
        form = StudentForm()

    return render(request, 'attendance/register_student.html', {
        'form': form,
        'selected_school': selected_school
    })


@login_required
def select_school(request):
    user = request.user
    schools = School.objects.filter(user=user).order_by('id')  # 등록 순서대로

    color_list = ["#e6f2ff", "#e8f7e4", "#fff9e6", "#f0f4ff", "#f9f9f9"]
    school_colors = {}

    # 🔹 등급별 활성/비활성 처리: school_limit 이내만 활성
    for i, school in enumerate(schools):
        school_colors[school.id] = color_list[i % len(color_list)]
        school.is_active = i < user.school_limit  # Free=1개, Pro=5개

    if request.method == 'POST':
        selected_id = request.POST.get('school')
        return redirect(f"/attendance/?school={selected_id}")

    return render(request, 'attendance/select_school.html', {
        'schools': schools,
        "school_colors": school_colors,
    })

@login_required
def register_school(request):
    # ✅ 학교 생성 제한 (Free: 1개, Pro: 5개)
    current_school_count = School.objects.filter(user=request.user).count()
    if current_school_count >= request.user.school_limit:
        messages.warning(request, f"현재 등급에서는 최대 {request.user.school_limit}개의 학교만 등록할 수 있습니다. 더 많은 학교를 운영하시려면 Pro로 업그레이드하세요!", extra_tags='pricing_modal')
        return redirect('select_school')

    if request.method == 'POST':
        form = SchoolForm(request.POST)
        if form.is_valid():
            school = form.save(commit=False)
            school.user = request.user
            school.save()
            return redirect('select_school')
    else:
        form = SchoolForm()
    return render(request, 'attendance/register_school.html', {'form': form})

@login_required
def upload_students_excel(request):
    school_id = request.GET.get('school')
    selected_school = School.objects.filter(id=school_id, user=request.user).first()
    print(school_id)

    if not selected_school:
        return HttpResponse("유효한 학교 ID가 필요합니다.", status=400)

    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        wb = openpyxl.load_workbook(excel_file)
        # ✅ 학생 수 제한 확인 (학교당 80명)
        current_student_count = Student.objects.filter(school=selected_school).count()
        remaining_slots = request.user.student_limit_per_school - current_student_count

        if remaining_slots <= 0:
            messages.warning(request, f"학교당 최대 {request.user.student_limit_per_school}명까지만 등록 가능합니다. 무제한 등록을 위해 Pro로 업그레이드하세요!", extra_tags='pricing_modal')
            return redirect(f"/attendance/students/upload/?school={selected_school.id}")

        added_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if added_count >= remaining_slots:
                messages.warning(request, f"최대 인원(80명) 제한으로 인해 일부 학생만 등록되었습니다.")
                break

            department, grade, classroom, number, name, phone = row
            if not name: continue

            Student.objects.create(
                school=selected_school,
                department=department,
                grade=grade,
                classroom=classroom,
                number=number,
                name=name,
                phone=phone
            )
            added_count += 1

        messages.success(request, f"학생 {added_count}명의 엑셀 업로드가 완료되었습니다.")
        return redirect(f"/attendance/?school={selected_school.id}")

    return render(request, 'attendance/upload_excel.html', {
        'selected_school': selected_school
    })


@login_required
def download_student_excel_template(request):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "학생등록_양식"

    # 헤더 설정 (부서, 학년, 반, 번호, 이름, 전화번호)
    headers = ['부서', '학년', '반', '번호', '이름', '전화번호']
    sheet.append(headers)

    # 샘플 데이터 (선택 사항)
    sample_data = ['1부', 1, 1, 1, '홍길동', '01012345678']
    sheet.append(sample_data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=student_upload_template.xlsx'
    wb.save(response)
    return response


@csrf_exempt
def ajax_attendance_cancel(request, student_id):
    if request.method == 'POST':
        student = get_object_or_404(Student, id=student_id)
        
        # ✅ Get date from request body if available, else default to today
        data = json.loads(request.body or "{}")
        date_str = data.get('date')
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                target_date = timezone.now().date()
        else:
            target_date = timezone.now().date()
            
        user_settings, _ = Setting.objects.get_or_create(user=student.school.user)
        send_sms = False
        sms_message = ""
        status_to_save = '취소'
        
        if user_settings.send_cancel_sms:
            sms_message = resolve_and_render_message(
                school=student.school,
                status='취소',
                student_name=student.name,
                settings_obj=user_settings,
            )
            send_sms = True
        else:
            status_to_save = '취소(문자x)'
            
        Attendance.objects.filter(student=student, date=target_date).update(status=status_to_save)
            
        return JsonResponse({'status': 'canceled', 'student': student.name, 'send_sms': send_sms, 'sms_message': sms_message})
    return JsonResponse({'status': 'invalid'})


from django.views.decorators.csrf import csrf_exempt

@login_required
def ajax_attendance_check(request, student_id):
    student = get_object_or_404(Student, id=student_id, school__user=request.user)

    if request.method == 'POST':
        data = json.loads(request.body or "{}")
        
        # ✅ Get date from request body if available, else default to today
        date_str = data.get('date')
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                target_date = timezone.localdate()
        else:
            target_date = timezone.localdate()
        status = data.get('status', '출석')
        program = data.get('program_name')
        absence_reason = (data.get('absence_reason') or '').strip()
        if status != '결석':
            absence_reason = ''

        existing_attendance = Attendance.objects.filter(student=student, date=target_date).first()
        if existing_attendance and existing_attendance.status in ['출석', '결석', '종료처리']:
            return JsonResponse({'status': 'already_checked'})

        user = student.school.user
        settings, created = Setting.objects.get_or_create(user=user)

        send_sms = False
        sms_message = ""

        if status == '출석':
            if settings.send_attendance_sms:
                sms_message = resolve_and_render_message(student.school, '출석', student.name, settings)
                send_sms = True
            else:
                status = '출석(문자x)'
        elif status == '지각':
            if settings.send_lateness_sms and settings.auto_send_lateness_sms:
                sms_message = resolve_and_render_message(student.school, '지각', student.name, settings)
                send_sms = True
            else:
                status = '지각(문자x)'
        elif status == '결석':
            if settings.send_absence_sms:
                sms_message = resolve_and_render_message(student.school, '결석', student.name, settings)
                send_sms = True
            else:
                status = '결석(문자x)'

        if existing_attendance:
            existing_attendance.status = status
            existing_attendance.absence_reason = absence_reason or None
            existing_attendance.program = program
            existing_attendance.created_at = timezone.now()
            existing_attendance.save(update_fields=['status', 'absence_reason', 'program', 'created_at'])
            attendance = existing_attendance
        else:
            attendance = Attendance.objects.create(
                student=student,
                status=status,
                absence_reason=absence_reason or None,
                program=program,
                date=target_date
            )

        created_time = timezone.localtime(attendance.created_at).strftime('%H:%M:%S')

        return JsonResponse({
            'status': 'success',
            'student': student.name,
            'phone': student.phone,
            'attendance_status': status,
            'absence_reason': attendance.absence_reason or '',
            'program_name': program,
            'created_at': created_time,
            'send_sms': send_sms,
            'sms_message': sms_message
        })

    return JsonResponse({'status': 'invalid_method'})



def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('attendance_list')  # 또는 적절한 페이지
    else:
        form = StudentForm(instance=student)

    return render(request, 'attendance/student_form.html', {
        'form': form,
        'student': student,
        'selected_school': student.school
        })


@login_required
def student_attendance_history(request, pk):
    student = get_object_or_404(Student, pk=pk, school__user=request.user)
    attendance_history = Attendance.objects.filter(student=student).order_by('-date', '-created_at')

    monthly_history = OrderedDict()
    for record in attendance_history:
        month_key = record.date.strftime('%Y-%m')
        if month_key not in monthly_history:
            monthly_history[month_key] = {
                'label': record.date.strftime('%Y년 %m월'),
                'records': []
            }
        monthly_history[month_key]['records'].append(record)

    return render(request, 'attendance/student_attendance_history.html', {
        'student': student,
        'selected_school': student.school,
        'monthly_history': list(monthly_history.values()),
    })

def student_create(request):
    school_id = request.GET.get('school')
    selected_school = School.objects.filter(id=school_id, user=request.user).first()

    if not selected_school:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': '학교를 찾을 수 없습니다.'}, status=400)
        return redirect('select_school')

    # ✅ 학생 수 제한 확인 (학교당 80명)
    current_student_count = Student.objects.filter(school=selected_school).count()
    if current_student_count >= request.user.student_limit_per_school:
        messages.warning(request, f"학교당 최대 {request.user.student_limit_per_school}명까지만 등록 가능합니다. 더 많은 학생 관리가 필요하시면 Pro로 업그레이드하세요!", extra_tags='pricing_modal')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': '인원 제한 도달 (80명). Pro 업그레이드 권장.', 'is_limit_error': True}, status=403)
        return redirect(f"/attendance/?school={selected_school.id}")

    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            import json
            from django.http import QueryDict
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'status': 'error', 'message': '잘못된 데이터 형식입니다.'}, status=400)

            post_data = QueryDict('', mutable=True)
            for key, value in data.items():
                post_data[key] = value
            form = StudentForm(post_data)
        else:
            form = StudentForm(request.POST)

        if form.is_valid():
            student = form.save(commit=False)
            student.school = selected_school
            student.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'student': {
                        'id': student.id,
                        'name': student.name,
                        'department': student.department,
                        'grade': student.grade,
                        'classroom': student.classroom,
                        'number': student.number,
                        'phone': student.phone
                    }
                })

            return redirect(f'/attendance/?school={selected_school.id}')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {field: errors[0] for field, errors in form.errors.items()}
                return JsonResponse({'status': 'error', 'errors': errors}, status=400)
    else:
        form = StudentForm()

    return render(request, 'attendance/student_form.html', {
        'form': form,
        'selected_school': selected_school
    })
    
@login_required(login_url="/accounts/login/")
def attendance_list(request):
    # ✅ 혹시라도 비로그인으로 들어오면 즉시 로그인으로
    if not request.user.is_authenticated:
        return redirect("/accounts/login/?next=/attendance/list/")

    user_id = request.user.id  # ✅ 정수 FK 안정화

    # ✅ 로그인 사용자의 학교 목록 (FK는 user_id로 필터링)
    schools = School.objects.filter(user_id=user_id)

    # ✅ GET 파라미터에서 선택된 학교 ID 가져오기
    selected_school_id = request.GET.get("school")

    if selected_school_id:
        selected_school = School.objects.filter(
            id=selected_school_id,
            user_id=user_id
        ).first()
    else:
        selected_school = schools.first()

    # 🔹 [보안] 비활성 학교 접근 차단
    if selected_school:
        active_school_ids = list(schools.order_by('id').values_list('id', flat=True)[:request.user.school_limit])
        if selected_school.id not in active_school_ids:
            from django.contrib import messages
            messages.warning(request, "해당 학교는 현재 비활성 상태입니다. Pro로 업그레이드하면 다시 이용할 수 있습니다.", extra_tags='pricing_modal')
            return redirect('select_school')

    if selected_school and selected_school.departments:
        department_options = [
            dept.strip()
            for dept in selected_school.departments.split(',')
            if dept.strip()
        ]
    else:
        department_options = ["1부", "2부", "3부"]

    # ✅ 선택된 학교에 해당하는 학생만 조회
    students = Student.objects.filter(school=selected_school) if selected_school else []

    today = timezone.now().date()
    daily_date_param = request.GET.get('daily_date')
    if daily_date_param:
        try:
            daily_target_date = datetime.strptime(daily_date_param, '%Y-%m-%d').date()
        except ValueError:
            daily_target_date = today
    else:
        daily_target_date = today

    daily_present_attendances = []
    daily_absent_attendances = []
    if selected_school:
        daily_attendance_qs = Attendance.objects.filter(
            date=daily_target_date,
            student__school=selected_school
        ).select_related('student').order_by('student__department', 'student__grade', 'student__classroom', 'student__number')

        for attendance in daily_attendance_qs:
            status = attendance.status or ""
            if status.startswith('결석'):
                daily_absent_attendances.append(attendance)
            elif status.startswith('출석') or status.startswith('지각') or status.startswith('종료처리'):
                daily_present_attendances.append(attendance)

    class_session_active = False
    if selected_school:
        class_session_active = AttendanceSession.objects.filter(
            school=selected_school,
            date=today,
            is_active=True
        ).exists()

    # ✅ (권장) 선택된 학교의 학생들만 해당 날짜 출석 가져오기
    #    학교가 없으면 빈 dict
    if selected_school:
        attendances_qs = Attendance.objects.filter(
            date=daily_target_date,
            student__school=selected_school
        )
    else:
        attendances_qs = Attendance.objects.none()

    attendances = {a.student_id: a for a in attendances_qs}

    # ✅ 부서별로 묶되, 학년-반-번호로 정렬
    department_groups = defaultdict(list)
    for student in students:
        department_groups[student.department].append(student)

    for dept, student_list in department_groups.items():
        department_groups[dept] = sorted(
            student_list,
            key=lambda s: (s.grade, s.classroom, s.number)
        )

    department_time_labels = {}
    if selected_school:
        for dept, time_info in selected_school.get_today_department_times(daily_target_date).items():
            if not time_info:
                continue
            start_time = time_info.get("start")
            end_time = time_info.get("end")
            if start_time and end_time:
                department_time_labels[dept] = f"{start_time}~{end_time}"

    # ✅ 부서 출력 순서 고정 (선택된 학교의 부서를 기준으로 표시)
    ordered_departments = OrderedDict()
    for dept in department_options:
        ordered_departments[dept] = department_groups.get(dept, [])

    user_settings, _ = Setting.objects.get_or_create(user=request.user)

    return render(request, "attendance/attendance_list.html", {
        "departments": ordered_departments,
        "attendances": attendances,
        "schools": schools,
        "selected_school": selected_school,
        "department_options": department_options,
        "department_time_labels": department_time_labels,
        "color_map_header": color_map_header,
        "class_session_active": class_session_active,
        "daily_target_date": daily_target_date,
        "daily_date_iso": daily_target_date.isoformat(),
        "daily_present_attendances": daily_present_attendances,
        "daily_absent_attendances": daily_absent_attendances,
        "active_dates_json": json.dumps([d.isoformat() for d in Attendance.objects.filter(student__school=selected_school).values_list('date', flat=True).distinct()]) if selected_school else '[]',
        "user_settings": user_settings,
    })

def attendance_check(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    today = timezone.now().date()
    # 이미 출석한 학생이면 중복 방지
    already_checked = Attendance.objects.filter(student=student, date=today).exists()
    if not already_checked:
        Attendance.objects.create(student=student)
        messages.success(request, f"{student.name}님 출석 완료!")
        # 👉 여기에서 문자 보내는 로직이 나중에 들어갈 부분!
    else:
        messages.warning(request, f"{student.name}님은 이미 출석했습니다.")
    return redirect('attendance_list')


@login_required
@require_POST
def move_students(request):
    try:
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        target_department = data.get('target_department')

        if not student_ids or not target_department:
            return JsonResponse({'status': 'error', 'message': '필수 정보가 누락되었습니다.'}, status=400)

        # Ensure the user is only moving students within their own schools
        students_to_move = Student.objects.filter(id__in=student_ids, school__user=request.user)
        
        if len(student_ids) != students_to_move.count():
            return JsonResponse({'status': 'error', 'message': '권한이 없거나 존재하지 않는 학생이 포함되어 있습니다.'}, status=403)

        updated_count = students_to_move.update(department=target_department)

        return JsonResponse({'status': 'success', 'message': f'{updated_count}명의 학생이 {target_department}로 이동되었습니다.'})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '잘못된 데이터 형식입니다.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
@require_POST
def end_class(request):
    try:
        data = json.loads(request.body)
        department_name = data.get('department_name')
        school_id = data.get('school_id')
        
        # ✅ Get date from request body if available, else default to today
        date_str = data.get('date')
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                target_date = timezone.now().date()
        else:
            target_date = timezone.now().date()

        if not department_name or not school_id:
            return JsonResponse({'status': 'error', 'message': '필수 정보가 누락되었습니다.'}, status=400)

        school = get_object_or_404(School, id=school_id, user=request.user)
        user_settings, _ = Setting.objects.get_or_create(user=request.user)

        attendances_to_end = Attendance.objects.filter(
            student__school=school,
            student__department=department_name,
            date=target_date,
            status='출석'
        )

        if not attendances_to_end.exists():
            return JsonResponse({'status': 'info', 'message': '수업을 종료할 출석 상태의 학생이 없습니다.'})

        student_count = attendances_to_end.count()
        phone_numbers = [att.student.phone for att in attendances_to_end if att.student.phone]

        attendances_to_end.update(status='종료처리')

        sms_uri = None
        if user_settings.send_class_end_sms and user_settings.auto_send_class_end_sms and phone_numbers:
            message = resolve_and_render_message(school, '종료처리', '', user_settings).strip()
            sms_uri = f"sms:{','.join(phone_numbers)}?body={message}"

        return JsonResponse({
            'status': 'success',
            'message': f'{department_name} 수업이 종료되었습니다. ({student_count}명 처리)',
            'sms_uri': sms_uri
        })

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': '잘못된 데이터 형식입니다.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@login_required
@require_POST
def ajax_update_setting(request):
    try:
        data = json.loads(request.body)
        if not field:
            return JsonResponse({'status': 'error', 'message': 'Field name is required.'}, status=400)

        # ✅ 보안: 자동화 설정은 Pro 유저만 변경 가능
        pro_only_fields = ['auto_send_lateness_sms', 'auto_send_class_end_sms']
        if field in pro_only_fields and not request.user.can_use_automation:
            return JsonResponse({
                'status': 'error', 
                'message': '자동화 설정은 Pro 멤버십 전용 기능입니다.'
            }, status=403)
            
        settings, _ = Setting.objects.get_or_create(user=request.user)
        
        if hasattr(settings, field):
            setattr(settings, field, value)
            settings.save()
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'message': f'Invalid field: {field}'}, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON.'}, status=400)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def pricing_page(request):
    import os
    context = {
        'PORTONE_SHOP_ID': os.getenv("PORTONE_SHOP_ID", "imp12345678")
    }
    return render(request, 'attendance/pricing.html', context)

def terms_page(request):
    """이용약관 페이지"""
    return render(request, 'attendance/terms.html')

def privacy_page(request):
    """개인정보처리방침 페이지"""
    return render(request, 'attendance/privacy.html')

def refund_page(request):
    """환불 정책 페이지"""
    return render(request, 'attendance/refund.html')
